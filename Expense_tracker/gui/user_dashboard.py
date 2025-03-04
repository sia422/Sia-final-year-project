import customtkinter as ctk
from tkinter import messagebox, filedialog
from tkcalendar import DateEntry
from services.expense_service import get_expenses, add_expense, update_expense, delete_expense
from services.income_service import get_incomes, get_total_income, add_income, update_income, delete_income
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook
from database.db import initialize_db  # Import the database initializer


class UserDashboard:
    def __init__(self, root, user_id, username):
        """
        Initializes the user dashboard with a professional and responsive design.
        """
        self.root = root
        self.user_id = user_id
        self.username = username

        # Set up the main window properties
        self.root.title(f"Income & Expense Tracker - {username}")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)

        # Configure grid layout for responsiveness
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)

        # Sidebar (Fixed width)
        self.sidebar = ctk.CTkFrame(self.root, width=250, corner_radius=8, fg_color="#2B2B2B")
        self.sidebar.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        # Header in Sidebar
        ctk.CTkLabel(
            self.sidebar,
            text=f"✨ Welcome, {username}",
            font=("Arial", 16, "bold"),
            text_color="white"
        ).pack(pady=(20, 10), padx=10)

        # Sidebar Buttons
        buttons = [
            ("🏠 Dashboard", self.show_dashboard),
            ("💰 My Expenses", self.show_expenses),
            ("💸 My Incomes", self.show_incomes),
            ("📊 Generate Reports", self.show_reports),
            ("🚪 Logout", self.logout)
        ]
        for text, command in buttons:
            ctk.CTkButton(
                self.sidebar,
                text=text,
                font=("Arial", 14),
                corner_radius=8,
                command=command,
                fg_color="#4A4A4A",
                hover_color="#333333"
            ).pack(fill="x", padx=10, pady=5)

        # Main Content Area
        self.main_content = ctk.CTkFrame(self.root, corner_radius=8, fg_color="#2B2B2B")
        self.main_content.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        # Load default view
        self.show_dashboard()

    def clear_main_content(self):
        """Clear all widgets in the main content area."""
        for widget in self.main_content.winfo_children():
            widget.destroy()

    def show_dashboard(self):
        """Display the dashboard with user analytics and charts."""
        self.clear_main_content()

        # Ensure the database is initialized
        initialize_db()

        # Fetch data
        expenses = get_expenses(self.user_id)
        total_expenses = sum(expense["amount"] for expense in expenses) if expenses else 0
        incomes = get_incomes(self.user_id)
        total_incomes = sum(income["amount"] for income in incomes) if incomes else 0

        # Dashboard Title
        ctk.CTkLabel(
            self.main_content,
            text="🏠 Dashboard Overview",
            font=("Arial", 24, "bold"),
            text_color="white"
        ).pack(pady=(20, 10))

        # Analytics Section with Professional Cards
        analytics_frame = ctk.CTkFrame(self.main_content, corner_radius=8, fg_color="#3B3B3B")
        analytics_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # Total Expenses Card
        expenses_card = ctk.CTkFrame(analytics_frame, corner_radius=8, fg_color="#FF4500")
        expenses_card.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        ctk.CTkLabel(
            expenses_card,
            text="Total Expenses",
            font=("Arial", 18, "bold"),
            text_color="white"
        ).pack(pady=(10, 5), padx=10)
        ctk.CTkLabel(
            expenses_card,
            text=f"NLe {total_expenses:.2f}",  # Changed currency to NLe
            font=("Arial", 24, "bold"),
            text_color="white"
        ).pack(pady=(5, 10), padx=10)

        # Total Incomes Card
        incomes_card = ctk.CTkFrame(analytics_frame, corner_radius=8, fg_color="#2E8B57")
        incomes_card.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        ctk.CTkLabel(
            incomes_card,
            text="Total Incomes",
            font=("Arial", 18, "bold"),
            text_color="white"
        ).pack(pady=(10, 5), padx=10)
        ctk.CTkLabel(
            incomes_card,
            text=f"NLe {total_incomes:.2f}",  # Changed currency to NLe
            font=("Arial", 24, "bold"),
            text_color="white"
        ).pack(pady=(5, 10), padx=10)

        # Net Balance Card
        net_balance = total_incomes - total_expenses
        balance_color = "#3CB371" if net_balance >= 0 else "#FF4500"
        balance_card = ctk.CTkFrame(analytics_frame, corner_radius=8, fg_color=balance_color)
        balance_card.grid(row=0, column=2, padx=10, pady=10, sticky="nsew")
        ctk.CTkLabel(
            balance_card,
            text="Net Balance",
            font=("Arial", 18, "bold"),
            text_color="white"
        ).pack(pady=(10, 5), padx=10)
        ctk.CTkLabel(
            balance_card,
            text=f"NLe {net_balance:.2f}",  # Changed currency to NLe
            font=("Arial", 24, "bold"),
            text_color="white"
        ).pack(pady=(5, 10), padx=10)

        # Configure grid weights for analytics cards
        analytics_frame.grid_columnconfigure((0, 1, 2), weight=1)

        # Chart Section
        chart_frame = ctk.CTkFrame(self.main_content, corner_radius=8, fg_color="#3B3B3B")
        chart_frame.pack(fill="both", expand=True, padx=20, pady=10)
        self.plot_pie_chart(total_expenses, total_incomes)

    def plot_pie_chart(self, total_expenses, total_incomes):
        """Plot a large, centered pie chart showing income vs. expenses."""
        if total_expenses == 0 and total_incomes == 0:
            ctk.CTkLabel(
                self.main_content,
                text="No data available to display charts.",
                font=("Arial", 16),
                text_color="white"
            ).pack(pady=10)
            return

        # Create the pie chart
        fig, ax = plt.subplots(figsize=(6, 6))  # Larger size for better visibility
        labels = ['Expenses', 'Incomes']
        sizes = [total_expenses, total_incomes]
        ax.pie(
            sizes,
            labels=labels,
            autopct='%1.1f%%',
            startangle=90,
            colors=["#FF4500", "#2E8B57"],
            wedgeprops={'edgecolor': 'black'}
        )
        ax.axis('equal')  # Equal aspect ratio ensures the pie chart is circular.
        ax.set_title("Income vs. Expenses", fontsize=18, color="white")

        # Embed the chart into the dashboard
        canvas = FigureCanvasTkAgg(fig, master=self.main_content)
        canvas.draw()
        canvas.get_tk_widget().pack(side="top", fill="both", expand=True, padx=10, pady=10)

    def show_expenses(self):
        """Display the user's expenses with CRUD options."""
        self.clear_main_content()

        # Ensure the database is initialized
        initialize_db()

        # Expense Form Section
        form_frame = ctk.CTkFrame(self.main_content, corner_radius=8, fg_color="#3B3B3B")
        form_frame.pack(fill="both", expand=False, padx=20, pady=10)
        ctk.CTkLabel(
            form_frame,
            text="Add New Expense",
            font=("Arial", 18, "bold"),
            text_color="white"
        ).pack(pady=(10, 5), padx=10)

        # Inputs for Expense
        self.entry_title = ctk.CTkEntry(form_frame, placeholder_text="Enter expense title", font=("Arial", 14), corner_radius=8)
        self.entry_amount = ctk.CTkEntry(form_frame, placeholder_text="Enter expense amount", font=("Arial", 14), corner_radius=8)
        self.combo_category = ctk.CTkOptionMenu(
            form_frame,
            values=["Food", "Transport", "Entertainment", "Other"],
            font=("Arial", 14),
            dropdown_fg_color="#3B3B3B",
            button_color="#4A4A4A",
            button_hover_color="#333333"
        )
        self.entry_date = DateEntry(
            form_frame,
            selectmode="day",
            date_pattern="yyyy-mm-dd",
            background="#2B2B2B",
            foreground="white",
            normalbackground="#3B3B3B",
            normalforeground="white",
            headersbackground="#4A4A4A",
            headersforeground="white",
            bordercolor="#4A4A4A",
            selectbackground="#FFA500",
            selectforeground="black"
        )

        # Title Input
        ctk.CTkLabel(form_frame, text="Title:", font=("Arial", 14), text_color="white").pack(pady=(10, 5), padx=10, anchor="w")
        self.entry_title.pack(fill="x", padx=20, pady=5)

        # Amount Input
        ctk.CTkLabel(form_frame, text="Amount:", font=("Arial", 14), text_color="white").pack(pady=(10, 5), padx=10, anchor="w")
        self.entry_amount.pack(fill="x", padx=20, pady=5)

        # Category Dropdown
        ctk.CTkLabel(form_frame, text="Category:", font=("Arial", 14), text_color="white").pack(pady=(10, 5), padx=10, anchor="w")
        self.combo_category.pack(fill="x", padx=20, pady=5)

        # Date Input (Embedded Calendar)
        ctk.CTkLabel(form_frame, text="Date:", font=("Arial", 14), text_color="white").pack(pady=(10, 5), padx=10, anchor="w")
        self.entry_date.pack(fill="x", padx=20, pady=5)

        # Add Expense Button
        ctk.CTkButton(
            form_frame,
            text="➕ Add Expense",
            font=("Arial", 14),
            command=self.add_expense,
            fg_color="#2E8B57",
            hover_color="#3CB371",
            corner_radius=8
        ).pack(pady=10)

        # Export Expenses to Excel Button
        ctk.CTkButton(
            form_frame,
            text="Export Expenses to Excel",
            font=("Arial", 14),
            command=self.export_expenses_to_excel,
            fg_color="#FFA500",
            hover_color="#FF8C00",
            corner_radius=8
        ).pack(pady=10)

        # Expense List Section
        list_frame = ctk.CTkScrollableFrame(self.main_content, corner_radius=8, fg_color="#3B3B3B")
        list_frame.pack(fill="both", expand=True, padx=20, pady=10)
        columns = ("ID", "Title", "Amount", "Category", "Date", "Actions")
        self.load_expenses(list_frame, columns)

    def load_expenses(self, parent_frame, columns):
        """Load and display all expenses for the current user with inline buttons."""
        header_frame = ctk.CTkFrame(parent_frame, corner_radius=8, fg_color="#3B3B3B")
        header_frame.pack(fill="x", pady=(10, 5))
        for idx, col in enumerate(columns):
            ctk.CTkLabel(
                header_frame,
                text=col,
                font=("Arial", 14, "bold"),
                text_color="white"
            ).grid(row=0, column=idx, padx=10, pady=5, sticky="w")

        expenses = get_expenses(self.user_id)
        for idx, expense in enumerate(expenses):
            row_frame = ctk.CTkFrame(parent_frame, corner_radius=6, fg_color="#4A4A4A")
            row_frame.pack(fill="x", pady=5, padx=10)

            ctk.CTkLabel(
                row_frame,
                text=str(expense["id"]),
                font=("Arial", 14),
                text_color="white"
            ).grid(row=0, column=0, padx=10, pady=5, sticky="w")

            ctk.CTkLabel(
                row_frame,
                text=expense["title"],
                font=("Arial", 14),
                text_color="white"
            ).grid(row=0, column=1, padx=10, pady=5, sticky="w")

            ctk.CTkLabel(
                row_frame,
                text=f"NLe {expense['amount']:.2f}",  # Changed currency to NLe
                font=("Arial", 14),
                text_color="white"
            ).grid(row=0, column=2, padx=10, pady=5, sticky="w")

            ctk.CTkLabel(
                row_frame,
                text=expense["category"],
                font=("Arial", 14),
                text_color="white"
            ).grid(row=0, column=3, padx=10, pady=5, sticky="w")

            ctk.CTkLabel(
                row_frame,
                text=expense["date"],
                font=("Arial", 14),
                text_color="white"
            ).grid(row=0, column=4, padx=10, pady=5, sticky="w")

            # Inline Buttons for Edit and Delete
            edit_button = ctk.CTkButton(
                row_frame,
                text="📝",
                font=("Arial", 12),
                width=30,
                command=lambda id=expense["id"]: self.edit_expense(id),
                fg_color="#FFA500",
                hover_color="#FF8C00"
            )
            edit_button.grid(row=0, column=5, padx=5, pady=5)

            delete_button = ctk.CTkButton(
                row_frame,
                text="🗑️",
                font=("Arial", 12),
                width=30,
                command=lambda id=expense["id"]: self.delete_expense(id),
                fg_color="#FF4500",
                hover_color="#FF0000"
            )
            delete_button.grid(row=0, column=6, padx=5, pady=5)

    def export_expenses_to_excel(self):
        """Export expenses to an Excel file."""
        try:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if not file_path:
                return

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Expenses"

            # Write headers
            headers = ["ID", "Title", "Amount (NLe)", "Category", "Date"]
            sheet.append(headers)

            # Write data rows
            expenses = get_expenses(self.user_id)
            for expense in expenses:
                sheet.append([
                    expense["id"],
                    expense["title"],
                    f"{expense['amount']:.2f}",  # Format amount to two decimal places
                    expense["category"],
                    expense["date"]
                ])

            # Save the workbook
            workbook.save(file_path)
            messagebox.showinfo("Success", "Expenses exported to Excel successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export expenses: {str(e)}")

    def add_expense(self):
        """Add a new expense to the database and refresh the list."""
        title = self.entry_title.get().strip()
        amount = self.entry_amount.get().strip()
        category = self.combo_category.get().strip()
        date = self.entry_date.get().strip()

        if not title or not amount or not category or not date:
            messagebox.showwarning("Input Error", "❌ Please fill out all fields.")
            return

        if not amount.replace('.', '', 1).isdigit():
            messagebox.showwarning("Input Error", "❌ Please enter a valid amount (numeric value).")
            return

        try:
            if add_expense(self.user_id, title, float(amount), "NLe", category, date):  # Changed currency to NLe
                messagebox.showinfo("Success", "✅ Expense added successfully!")
                self.clear_form()
                self.show_expenses()
            else:
                messagebox.showerror("Error", "❌ Failed to add expense.")
        except Exception as e:
            messagebox.showerror("Error", f"❌ An error occurred: {str(e)}")

    def clear_form(self):
        """Clear the input fields in the expense form."""
        if hasattr(self, "entry_title"):
            self.entry_title.delete(0, ctk.END)
        if hasattr(self, "entry_amount"):
            self.entry_amount.delete(0, ctk.END)
        if hasattr(self, "combo_category"):
            self.combo_category.set("")
        if hasattr(self, "entry_date"):
            self.entry_date.set_date(None)  # Reset the DateEntry widget

    def edit_expense(self, expense_id):
        """Edit an existing expense."""
        title = self._get_input("Enter new title:", "Edit Expense")
        if not title:
            return

        amount = self._get_numeric_input("Enter new amount:", "Edit Expense")
        if amount is None:
            return

        category = self._get_input("Enter new category name:", "Edit Expense")
        if not category:
            return

        date = self._get_input("Enter new date (YYYY-MM-DD):", "Edit Expense")
        if not date:
            return

        try:
            if update_expense(expense_id, title, float(amount), "NLe", category, date):  # Changed currency to NLe
                messagebox.showinfo("Success", "✅ Expense updated successfully!")
                self.show_expenses()
            else:
                messagebox.showerror("Error", "❌ Failed to update expense.")
        except Exception as e:
            messagebox.showerror("Error", f"❌ An error occurred: {str(e)}")

    def delete_expense(self, expense_id):
        """Delete an existing expense."""
        confirm = messagebox.askyesno("Confirm", "Are you sure you want to delete this expense?")
        if confirm:
            try:
                if delete_expense(expense_id):
                    messagebox.showinfo("Success", "✅ Expense deleted successfully!")
                    self.show_expenses()
                else:
                    messagebox.showerror("Error", "❌ Failed to delete expense.")
            except Exception as e:
                messagebox.showerror("Error", f"❌ An error occurred: {str(e)}")

    def show_incomes(self):
        """Display the user's incomes with CRUD options."""
        self.clear_main_content()

        # Ensure the database is initialized
        initialize_db()

        # Income Form Section
        form_frame = ctk.CTkFrame(self.main_content, corner_radius=8, fg_color="#3B3B3B")
        form_frame.pack(fill="both", expand=False, padx=20, pady=10)
        ctk.CTkLabel(
            form_frame,
            text="Add New Income",
            font=("Arial", 18, "bold"),
            text_color="white"
        ).pack(pady=(10, 5), padx=10)

        # Inputs for Income
        self.entry_income_source = ctk.CTkEntry(form_frame, placeholder_text="Enter income source", font=("Arial", 14), corner_radius=8)
        self.entry_income_amount = ctk.CTkEntry(form_frame, placeholder_text="Enter income amount", font=("Arial", 14), corner_radius=8)
        self.entry_income_date = DateEntry(
            form_frame,
            selectmode="day",
            date_pattern="yyyy-mm-dd",
            background="#2B2B2B",
            foreground="white",
            normalbackground="#3B3B3B",
            normalforeground="white",
            headersbackground="#4A4A4A",
            headersforeground="white",
            bordercolor="#4A4A4A",
            selectbackground="#FFA500",
            selectforeground="black"
        )

        # Source Input
        ctk.CTkLabel(form_frame, text="Source:", font=("Arial", 14), text_color="white").pack(pady=(10, 5), padx=10, anchor="w")
        self.entry_income_source.pack(fill="x", padx=20, pady=5)

        # Amount Input
        ctk.CTkLabel(form_frame, text="Amount:", font=("Arial", 14), text_color="white").pack(pady=(10, 5), padx=10, anchor="w")
        self.entry_income_amount.pack(fill="x", padx=20, pady=5)

        # Date Input (Embedded Calendar)
        ctk.CTkLabel(form_frame, text="Date:", font=("Arial", 14), text_color="white").pack(pady=(10, 5), padx=10, anchor="w")
        self.entry_income_date.pack(fill="x", padx=20, pady=5)

        # Add Income Button
        ctk.CTkButton(
            form_frame,
            text="➕ Add Income",
            font=("Arial", 14),
            command=self.add_income,
            fg_color="#2E8B57",
            hover_color="#3CB371",
            corner_radius=8
        ).pack(pady=10)

        # Export Incomes to Excel Button
        ctk.CTkButton(
            form_frame,
            text="Export Incomes to Excel",
            font=("Arial", 14),
            command=self.export_incomes_to_excel,
            fg_color="#FFA500",
            hover_color="#FF8C00",
            corner_radius=8
        ).pack(pady=10)

        # Income List Section
        list_frame = ctk.CTkScrollableFrame(self.main_content, corner_radius=8, fg_color="#3B3B3B")
        list_frame.pack(fill="both", expand=True, padx=20, pady=10)
        columns = ("ID", "Source", "Amount", "Date", "Actions")
        self.load_incomes(list_frame, columns)

    def load_incomes(self, parent_frame, columns):
        """Load and display all incomes for the current user with inline buttons."""
        header_frame = ctk.CTkFrame(parent_frame, corner_radius=8, fg_color="#3B3B3B")
        header_frame.pack(fill="x", pady=(10, 5))
        for idx, col in enumerate(columns):
            ctk.CTkLabel(
                header_frame,
                text=col,
                font=("Arial", 14, "bold"),
                text_color="white"
            ).grid(row=0, column=idx, padx=10, pady=5, sticky="w")

        incomes = get_incomes(self.user_id)
        for idx, income in enumerate(incomes):
            row_frame = ctk.CTkFrame(parent_frame, corner_radius=6, fg_color="#4A4A4A")
            row_frame.pack(fill="x", pady=5, padx=10)

            ctk.CTkLabel(
                row_frame,
                text=str(income["id"]),
                font=("Arial", 14),
                text_color="white"
            ).grid(row=0, column=0, padx=10, pady=5, sticky="w")

            ctk.CTkLabel(
                row_frame,
                text=income["source"],
                font=("Arial", 14),
                text_color="white"
            ).grid(row=0, column=1, padx=10, pady=5, sticky="w")

            ctk.CTkLabel(
                row_frame,
                text=f"NLe {income['amount']:.2f}",  # Changed currency to NLe
                font=("Arial", 14),
                text_color="white"
            ).grid(row=0, column=2, padx=10, pady=5, sticky="w")

            ctk.CTkLabel(
                row_frame,
                text=income["date"],
                font=("Arial", 14),
                text_color="white"
            ).grid(row=0, column=3, padx=10, pady=5, sticky="w")

            # Inline Buttons for Edit and Delete
            edit_button = ctk.CTkButton(
                row_frame,
                text="📝",
                font=("Arial", 12),
                width=30,
                command=lambda id=income["id"]: self.edit_income(id),
                fg_color="#FFA500",
                hover_color="#FF8C00"
            )
            edit_button.grid(row=0, column=4, padx=5, pady=5)

            delete_button = ctk.CTkButton(
                row_frame,
                text="🗑️",
                font=("Arial", 12),
                width=30,
                command=lambda id=income["id"]: self.delete_income(id),
                fg_color="#FF4500",
                hover_color="#FF0000"
            )
            delete_button.grid(row=0, column=5, padx=5, pady=5)

    def export_incomes_to_excel(self):
        """Export incomes to an Excel file."""
        try:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if not file_path:
                return

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Incomes"

            # Write headers
            headers = ["ID", "Source", "Amount (NLe)", "Date"]
            sheet.append(headers)

            # Write data rows
            incomes = get_incomes(self.user_id)
            for income in incomes:
                sheet.append([
                    income["id"],
                    income["source"],
                    f"{income['amount']:.2f}",  # Format amount to two decimal places
                    income["date"]
                ])

            # Save the workbook
            workbook.save(file_path)
            messagebox.showinfo("Success", "Incomes exported to Excel successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export incomes: {str(e)}")

    def add_income(self):
        """Add a new income to the database and refresh the list."""
        source = self.entry_income_source.get().strip()
        amount = self.entry_income_amount.get().strip()
        date = self.entry_income_date.get().strip()

        if not source or not amount or not date:
            messagebox.showwarning("Input Error", "❌ Please fill out all fields.")
            return

        if not amount.replace('.', '', 1).isdigit():
            messagebox.showwarning("Input Error", "❌ Please enter a valid amount (numeric value).")
            return

        try:
            if add_income(self.user_id, source, float(amount), date):
                messagebox.showinfo("Success", "✅ Income added successfully!")
                self.clear_form()
                self.show_incomes()
            else:
                messagebox.showerror("Error", "❌ Failed to add income.")
        except Exception as e:
            messagebox.showerror("Error", f"❌ An error occurred: {str(e)}")

    def edit_income(self, income_id):
        """Edit an existing income."""
        source = self._get_input("Enter new source:", "Edit Income")
        if not source:
            return

        amount = self._get_numeric_input("Enter new amount:", "Edit Income")
        if amount is None:
            return

        date = self._get_input("Enter new date (YYYY-MM-DD):", "Edit Income")
        if not date:
            return

        try:
            if update_income(income_id, source, float(amount), date):
                messagebox.showinfo("Success", "✅ Income updated successfully!")
                self.show_incomes()
            else:
                messagebox.showerror("Error", "❌ Failed to update income.")
        except Exception as e:
            messagebox.showerror("Error", f"❌ An error occurred: {str(e)}")

    def delete_income(self, income_id):
        """Delete an existing income."""
        confirm = messagebox.askyesno("Confirm", "Are you sure you want to delete this income?")
        if confirm:
            try:
                if delete_income(income_id):
                    messagebox.showinfo("Success", "✅ Income deleted successfully!")
                    self.show_incomes()
                else:
                    messagebox.showerror("Error", "❌ Failed to delete income.")
            except Exception as e:
                messagebox.showerror("Error", f"❌ An error occurred: {str(e)}")

    def show_reports(self):
        """Generate and display reports with charts."""
        self.clear_main_content()

        # Ensure the database is initialized
        initialize_db()

        # Reports Title
        ctk.CTkLabel(
            self.main_content,
            text="📊 Reports",
            font=("Arial", 24, "bold"),
            text_color="white"
        ).pack(pady=(20, 10))

        # Export Combined Report Button
        ctk.CTkButton(
            self.main_content,
            text="Export Combined Report to Excel",
            font=("Arial", 14),
            command=self.export_combined_report,
            fg_color="#FFA500",
            hover_color="#FF8C00",
            corner_radius=8
        ).pack(pady=10)

        # Display Charts
        expenses = get_expenses(self.user_id)
        total_expenses = sum(expense["amount"] for expense in expenses) if expenses else 0
        incomes = get_incomes(self.user_id)
        total_incomes = sum(income["amount"] for income in incomes) if incomes else 0

        if total_expenses == 0 and total_incomes == 0:
            ctk.CTkLabel(
                self.main_content,
                text="No data available to display charts.",
                font=("Arial", 16),
                text_color="white"
            ).pack(pady=10)
            return

        # Plot Income vs. Expenses Pie Chart
        fig, ax = plt.subplots(figsize=(6, 6))  # Larger pie chart
        labels = ['Expenses', 'Incomes']
        sizes = [total_expenses, total_incomes]
        ax.pie(
            sizes,
            labels=labels,
            autopct='%1.1f%%',
            startangle=90,
            colors=["#FF4500", "#2E8B57"],
            wedgeprops={'edgecolor': 'black'}
        )
        ax.axis('equal')  # Equal aspect ratio ensures the pie chart is circular.
        ax.set_title("Income vs. Expenses", fontsize=18, color="white")

        # Embed the chart into the dashboard
        canvas = FigureCanvasTkAgg(fig, master=self.main_content)
        canvas.draw()
        canvas.get_tk_widget().pack(side="top", fill="both", expand=True, padx=10, pady=10)

    def export_combined_report(self):
        """Export a combined report of incomes and expenses to an Excel file."""
        try:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if not file_path:
                return

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Combined Report"

            # Write headers
            headers = ["ID", "Type", "Description", "Amount (NLe)", "Date"]
            sheet.append(headers)

            # Write expenses
            expenses = get_expenses(self.user_id)
            for expense in expenses:
                sheet.append([
                    expense["id"],
                    "Expense",
                    expense["title"],
                    f"{expense['amount']:.2f}",  # Format amount to two decimal places
                    expense["date"]
                ])

            # Write incomes
            incomes = get_incomes(self.user_id)
            for income in incomes:
                sheet.append([
                    income["id"],
                    "Income",
                    income["source"],
                    f"{income['amount']:.2f}",  # Format amount to two decimal places
                    income["date"]
                ])

            # Save the workbook
            workbook.save(file_path)
            messagebox.showinfo("Success", "Report exported to Excel successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export report: {str(e)}")

    def logout(self):
        """Log out the user and return to the authentication window."""
        self.root.destroy()
        from main import open_auth_window
        open_auth_window()

    def _get_input(self, prompt, title):
        """Helper method to get string input from the user."""
        return ctk.CTkInputDialog(text=prompt, title=title).get_input()

    def _get_numeric_input(self, prompt, title):
        """Helper method to get numeric input from the user."""
        value = self._get_input(prompt, title)
        if not value or not value.replace('.', '', 1).isdigit():
            messagebox.showerror("Error", "❌ Invalid input. Please enter a valid number.")
            return None
        return value


# Example usage (for testing purposes)
if __name__ == "__main__":
    app = ctk.CTk()
    ctk.set_appearance_mode("Dark")  # Dark mode only
    ctk.set_default_color_theme("blue")
    dashboard = UserDashboard(app, user_id=1, username="Mohamed")
    app.mainloop()